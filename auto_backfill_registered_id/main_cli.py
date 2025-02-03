# %%
# %load_ext autoreload
# %autoreload 2

from collections import Counter
from copy import deepcopy
from pathlib import Path

import numpy as np
import pandas as pd
from modules.check import check_dir, check_xlsx
from modules.utils import (add_total_sum, col_strip, copy_as_index,
                           load_config, read_ws, sync_catalog_value)
from openpyxl import load_workbook
from rich.console import Console
from rich.traceback import install

install() # debug
console = Console(record=True)
# -----------------------------------------------------------------------------/

# %%
# 保留原始資料格式
{
    "採購序號": "04",
    "原始序號": "0",
    "箱號": "0",
    "登錄號": "0",
}

# 相等檢查
cata_match2purc: dict = {
    '書名': "string",
    'ISBN': "string",
}

# replace to 交貨清單
cata_rp2_purc: dict = {
    '箱號': "Int64",
    '登錄號': "Int64",
    '分類號': "string",
    '館藏地代碼': "string",
    '資料類型/特藏號': "string",
    '出版年': "string",
    '部冊號': "string", # 加在書名後面
    'ISBN': "string",
}

# Load config
config_path: Path = Path(__file__).parent.joinpath("path.toml")
config = load_config(config_path, console)

# %% [markdown]
# ### class attrs

# %%
# 交貨清單
purchasing_wb: Path = Path(config["purchasing"]) # attr: wb_path
check_xlsx(purchasing_wb.resolve(), "交貨清單")

purchasing_st_rowidx: int = 2
purchasing_colnames: set = set([
    '序號', '採購序號', '原始序號', '箱號', '登錄號', '書名', '作者', '出版社', '出版年', 'ISBN',
    '館藏地代碼', '資料類型/特藏號', '定價', '折扣價', '總冊數', '數量', '小計',
    '主題分類', '得獎/推薦1', '得獎/推薦2', '套書/複本', '書單來源', '分類號', '冊數', '單位',
])
purchasing_dtype: dict = {
    '採購序號': "Int64", # Int64 可以處理 NAN, int64 不行
    '書名': "string",
    '出版年': "string",
    'ISBN': "string",
    '館藏地代碼': "string",
    '資料類型/特藏號': "string",
    '定價': "Int64",
    '折扣價': "Float64",
    '總冊數': "Int64",
    '數量': "Int64",
    '小計': "Float64",
    '分類號': "string",
    '冊數': "Int64",
}
purchasing_colalias: dict = {
    "採購序號": '採購序號',
    "書名": '書名',
    "館藏地代碼": '館藏地代碼',
    "資料類型/特藏號": '資料類型/特藏號',
    "出版年": '出版年',
    "ISBN": 'ISBN',
}
purchasing_df: pd.DataFrame

# 編目箱單
catalog_wb: Path = Path(config["catalog"]) # attr: wb_path
check_xlsx(catalog_wb.resolve(), "編目箱單")

catalog_st_rowidx: int = 1
catalog_colnames: set = set([
    '編目員', '箱號', '序號', '原序號', '登錄號', '類型', '分類號', '作者號', '年代', '部冊號', '書目',
    '舊登錄號', 'F10', 'F11', 'F12', 'F13', 'F14', 'F15', 'F16', 'F17', 'F18', 'F19 裝訂', 'F20 010d',
    'F21 805購價', 'F22 805Note', 'F23 805Attach', 'F24 805ISBN', 'F25 805FLDY', 'F26 805LOC',
    'F27 805ISSNOTE', 'F28 805CallNo', 'F29 805d', 'F30 681a', 'F31 681v', 'F32', 'F33', 'F34', 'F35', 'F36',
])
catalog_dtype: dict = {
    '箱號': "Int64",
    '原序號': "Int64", # '採購序號'
    '登錄號': "string",
    '分類號': "string",
    '部冊號': "string",
    '書目': "string", # '書名'
    'F10': "string", # '館藏地代碼'
    'F11': "string", # '資料類型/特藏號'
    'F14': "string", # '出版年'
    'F24 805ISBN': "string", # 'ISBN'
}
catalog_colalias: dict = {
    "採購序號": "原序號",
    "書名": '書目',
    "館藏地代碼": 'F10',
    "資料類型/特藏號": 'F11',
    "出版年": 'F14',
    "ISBN": 'F24 805ISBN',
}
catalog_df: pd.DataFrame

# New WorkBook
new_wb_dir: Path = Path(config["new_wb_dir"]) # attr: wb_path
check_dir(new_wb_dir.resolve())

new_wbname: str
new_wb: Path
new_df: pd.DataFrame

# %%
# Func: 偵測交貨清單
wb_names = load_workbook(purchasing_wb, data_only=True).sheetnames
purchasing_df = None

for name in wb_names:
    if name == "交貨清單":
        purchasing_df = read_ws(purchasing_wb, name, purchasing_dtype,
                                purchasing_st_rowidx, purchasing_colnames,
                                console)

# work sheet error
if purchasing_df is None:
    raise ValueError(f"找不到 [工作表]'交貨清單' 請確認輸入的檔案. File: '{purchasing_wb}'")

# strip string columns
col_strip(purchasing_df, purchasing_colalias["書名"])
col_strip(purchasing_df, purchasing_colalias["ISBN"])
raw_purc_df = deepcopy(purchasing_df)

# get valid filnal row
final_row = int(list(purchasing_df["採購序號"].dropna())[-1])
purchasing_df = purchasing_df.iloc[:final_row, :]

purchasing_df.reset_index(inplace=True)
console.print(len(purchasing_df.index), "\n")
purchasing_df

# %%
# Func: 偵測編目箱單 + 合併
wb_names = load_workbook(catalog_wb, data_only=True).sheetnames
catalog_df = None

for name in wb_names:
    tmp_df = read_ws(catalog_wb, name, catalog_dtype,
                     catalog_st_rowidx, catalog_colnames,
                     console)
    # concat work sheet
    if catalog_df is None:
        catalog_df = deepcopy(tmp_df)
    else:
        catalog_df = pd.concat([catalog_df, tmp_df])

# strip string columns
col_strip(catalog_df, catalog_colalias["書名"])
col_strip(catalog_df, catalog_colalias["ISBN"])
col_strip(catalog_df, "部冊號")

catalog_df.reset_index(inplace=True)
console.print(len(catalog_df.index), "\n")
catalog_df

# %%
# 交貨清單
copy_as_index(purchasing_df, purchasing_colalias["採購序號"], "purc_sn")
# 編目箱單
copy_as_index(catalog_df, catalog_colalias["採購序號"], "purc_sn")

purc_sn_pool = deepcopy(list(purchasing_df.index))

# %%
new_df = deepcopy(purchasing_df)
new_df.drop(new_df.index, inplace=True)

# Cases
handled_type_cnt = Counter()

for purc_sn in purc_sn_pool:
    
    console.rule()
    
    new_df, handled_type = \
            sync_catalog_value(new_df, purc_sn, cata_rp2_purc,
                               purchasing_df, purchasing_colalias,
                               catalog_df, catalog_colalias,
                               console)
    
    handled_type_cnt.update([handled_type])
    console.print(f":mage: 處理模式 : {handled_type}")
    console.line()
    
    # 如果編目已經處理完畢
    if len(catalog_df) == 0:
        console.rule()
        console.print("[green] '編目箱單' 已無資料\n")
        break

# %%
console.print(handled_type_cnt, "\n")
# add total sum
new_df = add_total_sum(new_df)
console.print(f"合計 (總冊數) : {new_df["總冊數"].values[-1]}")
console.print(f"合計 (小計) : {new_df["小計"].values[-1]}")
# save console log
new_logname = catalog_wb.stem.replace("(OK)", "").replace("編目箱單", "")
new_logname = f"(SR)處理紀錄{new_logname}.log"
new_log = new_wb_dir.joinpath(new_logname)
console.save_text(new_log)

# %%
catalog_df

# %%
purchasing_df

# %%
new_df.drop('index', axis=1, inplace=True)
set(new_df.columns).symmetric_difference(set(raw_purc_df.columns))

purchasing_df.drop('index', axis=1, inplace=True)
set(new_df.columns).symmetric_difference(set(raw_purc_df.columns))

# empty rows
empty_row = pd.Series([pd.NA] * len(new_df.columns), index=new_df.columns)
empty_df = pd.DataFrame([empty_row]*5)
set(empty_df.columns).symmetric_difference(set(raw_purc_df.columns))

new_df = pd.concat([new_df, pd.DataFrame([empty_row]*5), purchasing_df])
new_df["採購序號"] = new_df["採購序號"].astype("string")
new_df["採購序號"] = np.where(pd.isna(new_df["採購序號"]), new_df["採購序號"], new_df["採購序號"].str.zfill(4))

new_wbname = catalog_wb.stem.replace("(OK)", "").replace("編目箱單", "")
new_wbname = f"(SR)回填{new_wbname}{catalog_wb.suffix}"
new_wb = new_wb_dir.joinpath(new_wbname)
with open(new_wb, mode="wb") as f:
    new_df.to_excel(f, engine="openpyxl", sheet_name="交貨清單", index=False)
